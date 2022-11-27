import json
from openpyxl import load_workbook
from datetime import datetime

workbook = load_workbook(filename="Marksheet.xlsx", data_only=True)
sheet = workbook.active

grades = []
descriptor_levels = {
    1: "Not Achieved",
    2: "Elementary Achievement",
    3: "Moderate Achievement",
    4: "Adequate Achievement",
    5: "Substantial Achievement",
    6: "Meritorious Achievement",
    7: "Outstanding Achievement"
}

english_avg_mark = 0
xhosa_avg_mark = 0
life_skill_avg_mark = 0
social_science_avg_mark = 0 
math_avg_mark = 0
natural_science_avg_mark = 0 


def get_level(mark):
    if 0 <= mark <= 29:
        return 1
    elif 30 <= mark <= 39:
        return 2
    elif 40 <= mark <= 49:
        return 3
    elif 50 <= mark <= 59:
        return 4
    elif 60 <= mark <= 69:
        return 5
    elif 70 <= mark <= 79:
        return 6
    elif 80 <= mark <= 100:
        return 7

for item in sheet.iter_rows(min_row=49,
                           max_row=49,
                           min_col=4,
                           max_col=9,
                           values_only=True):
    english_avg_mark = item[0]
    xhosa_avg_mark = item[1]
    math_avg_mark = item[2]
    life_skill_avg_mark = item[3]
    natural_science_avg_mark = item[4]
    social_science_avg_mark = item[5]

for row in sheet.iter_rows(min_row=2,
                           max_row=47,
                           min_col=1,
                           max_col=12,
                           values_only=True):
    grade = {
        "full_name": row[0] + ", " + row[1],
        "english_mark": str(row[3]),
        "xhosa_mark": str(row[4]),
        "math_mark": str(row[5]),
        "life_skill_mark": str(row[6]),
        "natural_science_mark": str(row[7]),
        "social_science_mark": str(row[8]),
        "total_mark": str(row[9]),
        "avg_mark": str(round(row[10])),
        "passed": row[11],
        "english_level": str(get_level(row[3])),
        "xhosa_level": str(get_level(row[4])),
        "math_level": str(get_level(row[5])),
        "life_skill_level": str(get_level(row[6])),
        "natural_science_level": str(get_level(row[7])),
        "social_science_level": str(get_level(row[8])),
        "english_description": descriptor_levels.get(get_level(row[3]),""),
        "xhosa_description": descriptor_levels.get(get_level(row[4]),""),
        "math_description": descriptor_levels.get(get_level(row[5]),""),
        "life_skill_description": descriptor_levels.get(get_level(row[6]),""),
        "natural_science_description": descriptor_levels.get(get_level(row[7]),""),
        "social_science_description":descriptor_levels.get(get_level(row[8]),""),
        "english_avg_mark": str(round(english_avg_mark)),
        "xhosa_avg_mark": str(round(xhosa_avg_mark)),
        "math_avg_mark": str(round(math_avg_mark)),
        "life_skill_avg_mark": str(round(life_skill_avg_mark)),
        "natural_science_avg_mark": str(round(natural_science_avg_mark)),
        "social_science_avg_mark": str(round(social_science_avg_mark)),
    }

    grades.append(grade)

print(grades)
print(json.dumps(grades))